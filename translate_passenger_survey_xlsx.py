import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "pesquisa-de-satisfacao-do-passageiro-4-tri-2025.xlsx"
TARGET = ROOT / "pesquisa-de-satisfacao-do-passageiro-4-tri-2025-en.xlsx"
CACHE = ROOT / "translation_cache_pt_en.json"

ARGOS_BASE = ROOT / ".argos_runtime"
os.environ["XDG_DATA_HOME"] = str(ARGOS_BASE / "data")
os.environ["XDG_CONFIG_HOME"] = str(ARGOS_BASE / "config")
os.environ["XDG_CACHE_HOME"] = str(ARGOS_BASE / "cache")
os.environ["ARGOS_CHUNK_TYPE"] = "MINISBD"
os.environ["ARGOS_DEVICE_TYPE"] = "cpu"

import argostranslate.translate  # noqa: E402


SKIP_COLUMNS = {
    "4º Tri 2025": {3, 8, 9, 11, 12},  # airport, terminal, gate, airline, flight
}

MANUAL = {
    "4º Tri 2025": "Q4 2025",
    "LEGENDA": "LEGEND",
    "PROJETO": "PROJECT",
    "PESQUISA SATISFAÇÃO 2025": "2025 SATISFACTION SURVEY",
    "EMPRESA": "ORGANIZATION",
    "SECRETARIA NACIONAL DE AVIAÇÃO CIVIL - SAC": "NATIONAL CIVIL AVIATION SECRETARIAT - SAC",
    "ARQUIVO": "FILE",
    "QUESTIONÁRIO - (LEGENDA)": "QUESTIONNAIRE - (LEGEND)",
    "RESPOSTAS": "ANSWERS",
    "EMBARQUE e DESEMBARQUE": "BOARDING AND DISEMBARKATION",
    "EMBARQUE": "BOARDING",
    "DESEMBARQUE": "DISEMBARKATION",
    "Valores: 1 a 999999999": "Values: 1 to 999999999",
    "Valores: 1 a 999999999 ou A, B, C\nA nomenclatura pode mudar de aeroporto para aeroporto": "Values: 1 to 999999999 or A, B, C\nThe naming convention may vary from airport to airport",
    "dd/mm/aaaa": "dd/mm/yyyy",
    "ID  - ÁGORA": "ID - ÁGORA",
    "Data da pesquisa": "Survey date",
    "Hora Inicial da pesquisa": "Survey start time",
    "Hora final da pesquisa": "Survey end time",
    "Terminal de embarque - EMBARQUE\n\nTerminal de desembarque - DESEMBARQUE": "Boarding terminal - BOARDING\n\nDisembarkation terminal - DISEMBARKATION",
    "Tipo de Voo": "Flight type",
    "Processo": "Process",
    "Doméstico": "Domestic",
    "LOCALIZAÇÃO": "LOCATION",
    "PESSOA COM DEFICIÊNCIA": "PERSON WITH DISABILITY",
    "ACESSO": "ACCESS",
    "CHECK-IN": "CHECK-IN",
    "INSPEÇÃO DE SEGURANÇA": "SECURITY SCREENING",
    "ÓRGÃOS PÚBLICOS": "PUBLIC AGENCIES",
    "COMÉRCIO E SERVIÇOS": "RETAIL AND SERVICES",
    "AMBIENTE AEROPORTUÁRIO": "AIRPORT ENVIRONMENT",
    "RESTITUIÇÃO DE BAGAGENS": "BAGGAGE CLAIM",
    "SATISFAÇÃO GERAL": "OVERALL SATISFACTION",
    "PERFIL": "PROFILE",
    "CHAVE": "KEY",
    "PROCESSO": "PROCESS",
    "AEROPORTO": "AIRPORT",
    "DATA": "DATE",
    "MÊS": "MONTH",
    "INÍCIO COLETA": "SURVEY START",
    "FIM COLETA": "SURVEY END",
    "TERMINAL": "TERMINAL",
    "PORTÃO": "GATE",
    "TIPO DE VOO": "FLIGHT TYPE",
    "CIA AÉREA": "AIRLINE",
    "VOO": "FLIGHT",
    "CONEXÃO": "CONNECTION",
    "AQUISIÇÃO DA PASSAGEM": "TICKET PURCHASE",
    "MEIO DE AQUISIÇÃO DA PASSAGEM": "TICKET PURCHASE CHANNEL",
    "MEIO DE  TRANSPORTE PARA O AEROPORTO": "TRANSPORTATION TO THE AIRPORT",
    "POSSUI DEFICIÊNCIA": "HAS A DISABILITY",
    "UTILIZA RECURSO ASSISTIVO": "USES ASSISTIVE DEVICE",
    "SOLICITOU ASSISTÊNCIA ESPECIAL": "REQUESTED SPECIAL ASSISTANCE",
    "FORMA DE DESEMBARQUE UTILIZADA": "DISEMBARKATION METHOD USED",
    "AVALIAÇÃO DO MÉTODO DE DESEMBARQUE": "DISEMBARKATION METHOD RATING",
    "UTILIZOU O ESTACIONAMENTO?": "USED PARKING?",
    "FACILIDADE DE DESEMBARQUE NO MEIO-FIO": "EASE OF CURBSIDE DROP-OFF",
    "OPÇÕES DE TRANSPORTE ATÉ O AEROPORTO": "TRANSPORTATION OPTIONS TO THE AIRPORT",
    "FORMA DE CHECK-IN": "CHECK-IN METHOD",
    "PROCESSO DE CHECK IN": "CHECK-IN PROCESS",
    "TEMPO DE ESPERA NA FILA": "QUEUE WAITING TIME",
    "ORGANIZAÇÃO DAS FILAS": "QUEUE ORGANIZATION",
    "QUANTIDADE DE TOTENS AA": "NUMBER OF SELF-SERVICE KIOSKS",
    "QUANTIDADE DE BALCÕES": "NUMBER OF COUNTERS",
    "CORDIALIDADE DOS FUNCIONÁRIOS": "STAFF COURTESY",
    "TEMPO DE ATENDIMENTO": "SERVICE TIME",
    "PROCESSO DE AQUISIÇÃO DA PASSAGEM": "TICKET PURCHASE PROCESS",
    "ATENDIMENTO DA CIA. AÉREA": "AIRLINE SERVICE",
    "PROCESSO DE INSPEÇÃO DE SEGURANÇA": "SECURITY SCREENING PROCESS",
    "TEMPO DE ESPERA EM FILA": "QUEUE WAITING TIME",
    "CONTROLE MIGRATÓRIO": "IMMIGRATION CONTROL",
    "QUANTIDADE DE GUICHÊS": "NUMBER OF SERVICE WINDOWS",
    "CONTROLE ADUANEIRO": "CUSTOMS CONTROL",
    "ESTABELECIMENTOS DE ALIMENTAÇÃO": "FOOD AND BEVERAGE OUTLETS",
    "QUANTIDADE DE ESTABELECIMENTOS DE ALIMENTAÇÃO": "NUMBER OF FOOD AND BEVERAGE OUTLETS",
    "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS DE ALIMENTAÇÃO": "QUALITY AND VARIETY OF FOOD AND BEVERAGE OPTIONS",
    "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS DE ALIMENTAÇÃO": "PRICE-QUALITY RATIO OF FOOD AND BEVERAGE OUTLETS",
    "ESTABELECIMENTOS COMERCIAIS": "RETAIL OUTLETS",
    "QUANTIDADE DE ESTABELECIMENTOS COMERCIAIS": "NUMBER OF RETAIL OUTLETS",
    "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS COMERCIAIS": "QUALITY AND VARIETY OF RETAIL OPTIONS",
    "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS COMERCIAIS": "PRICE-QUALITY RATIO OF RETAIL OUTLETS",
    "ESTACIONAMENTO": "PARKING",
    "QUALIDADE DAS INSTALAÇÕES DE ESTACIONAMENTO": "QUALITY OF PARKING FACILITIES",
    "FACILIDADE PARA ENCONTRAR VAGAS": "EASE OF FINDING PARKING SPACES",
    "FACILIDADE DE ACESSO AO TERMINAL": "EASE OF ACCESS TO THE TERMINAL",
    "RELAÇÃO CUSTO X BENEFÍCIO": "VALUE FOR MONEY",
    "LOCALIZAÇÃO E DESLOCAMENTO": "LOCATION AND MOVEMENT",
    "SINALIZAÇÃO": "SIGNAGE",
    "DISPONIBILIDADE DE PAINÉIS DE INFORMAÇÕES DE VOO": "AVAILABILITY OF FLIGHT INFORMATION DISPLAYS",
    "ACESSIBILIDADE DO TERMINAL": "TERMINAL ACCESSIBILITY",
    "CONFORTO DA SALA DE EMBARQUE": "BOARDING LOUNGE COMFORT",
    "CONFORTO TÉRMICO": "THERMAL COMFORT",
    "CONFORTO ACÚSTICO": "ACOUSTIC COMFORT",
    "DISPONIBILIDADE DE ASSENTOS": "SEAT AVAILABILITY",
    "DISPONIBILIDADE DE ASSENTOS RESERVADOS": "AVAILABILITY OF RESERVED SEATS",
    "DISPONIBILIDADE DE TOMADAS": "POWER OUTLET AVAILABILITY",
    "INTERNET DISPONIBILIZADA PELO AEROPORTO": "INTERNET PROVIDED BY THE AIRPORT",
    "VELOCIDADE DE CONEXÃO": "CONNECTION SPEED",
    "FACILIDADE DE ACESSO À REDE": "EASE OF NETWORK ACCESS",
    "SANITÁRIOS": "RESTROOMS",
    "QUANTIDADE DE BANHEIROS": "NUMBER OF RESTROOMS",
    "LIMPEZA DOS BANHEIROS": "RESTROOM CLEANLINESS",
    "MANUTENÇÃO GERAL DOS SANITÁRIOS": "GENERAL RESTROOM MAINTENANCE",
    "LIMPEZA GERAL DO AEROPORTO": "OVERALL AIRPORT CLEANLINESS",
    "PROCESSO DE RESTITUIÇÃO DE BAGAGENS": "BAGGAGE CLAIM PROCESS",
    "FACILIDADE DE IDENTIFICAÇÃO DA ESTEIRA DE RESTITUIÇÃO": "EASE OF IDENTIFYING THE BAGGAGE CLAIM CAROUSEL",
    "TEMPO DE RESTITUIÇÃO": "BAGGAGE CLAIM TIME",
    "INTEGRIDADE DA BAGAGEM": "BAGGAGE INTEGRITY",
    "MOTIVO": "REASON",
    "NACIONALIDADE": "NATIONALITY",
    "GÊNERO": "GENDER",
    "IDADE": "AGE",
    "ESCOLARIDADE": "EDUCATION",
    "RENDA FAMILIAR": "HOUSEHOLD INCOME",
    "VIAJANDO SOZINHO": "TRAVELING ALONE",
    "NÚMERO DE ACOMPANHANTES": "NUMBER OF COMPANIONS",
    "MOTIVO DA VIAGEM": "TRIP PURPOSE",
    "QUANTIDADE DE VIAGENS NOS ÚLTIMOS 12 MESES": "NUMBER OF TRIPS IN THE LAST 12 MONTHS",
    "JÁ EMBARCOU/DESEMBARCOU ANTES NO AEROPORTO": "HAS BOARDED/DISEMBARKED AT THE AIRPORT BEFORE",
    "ANTECEDÊNCIA": "ARRIVAL LEAD TIME",
    "TEMPO DE ESPERA": "WAITING TIME",
    "COMENTÁRIOS ADICIONAIS": "ADDITIONAL COMMENTS",
    "Desembarque": "Disembarkation",
    "Embarque": "Boarding",
    "Internacional": "International",
    "Nacional": "Domestic",
    "Nenhuma": "None",
    "Brasileira": "Brazilian",
    "Estrangeira": "Foreign",
    "Masculino": "Male",
    "Feminino": "Female",
    "Não": "No",
    "Sim": "Yes",
    "NS/NR": "DK/NA",
    "Em conexão": "Connecting",
    "Embarcado no aeroporto": "Boarded at the airport",
    "Pelo passageiro": "By the passenger",
    "Por terceiros": "By third parties",
    "Diretamente com a cia. aérea pela internet ou aplicativo": "Directly with the airline via internet or app",
    "Diretamente com a cia. aérea pelo telefone": "Directly with the airline by phone",
    "Diretamente com a cia. aérea em loja física (inclusive no aeroporto)": "Directly with the airline at a physical store (including at the airport)",
    "Outros sites ou Agência de viagens": "Other websites or travel agency",
    "Outros meios": "Other means",
    "Outros": "Other",
    "Lazer": "Leisure",
    "Ônibus": "Bus",
    "Carona": "Ride",
    "Balcão": "Counter",
    "Aplicativos": "Apps",
    "Internet/Aplicativo": "Internet/App",
    "Totem AA": "Self-service airline kiosk",
    "Analfabeto": "No formal literacy",
    "Ensino fundamental": "Elementary education",
    "Ensino médio": "High school",
    "Superior": "Higher education",
    "Doutorado": "Doctorate",
    "Mestrado": "Master's degree",
    "Especialização de nível superior": "Graduate specialization",
    "Prefiro não informar": "Prefer not to say",
    "Até 1 salário mínimo": "Up to 1 minimum wage",
    "Entre 1 e 2 s.m": "Between 1 and 2 minimum wages",
    "Entre 2 e 4 s.m": "Between 2 and 4 minimum wages",
    "Entre 4 e 10 s.m": "Between 4 and 10 minimum wages",
    "Entre 10 e 20 s.m": "Between 10 and 20 minimum wages",
    "Mais de 20 s.m": "More than 20 minimum wages",
    "1 pessoa": "1 person",
    "2 pessoas": "2 people",
    "3 pessoas": "3 people",
    "4 ou mais pessoas": "4 or more people",
    "Trabalho": "Business",
    "Lazer e Trabalho": "Leisure and business",
    "Conforme necessário": "As applicable",
    "1 (Primeira viagem)": "1 (First trip)",
    "1 - primeira viagem": "1 - first trip",
    "2 a 3": "2 to 3",
    "4 a 5": "4 to 5",
    "6 a 10": "6 to 10",
    "Mais de 11": "More than 11",
    "30min a 1h": "30 min to 1 h",
    "1h a 1h30min": "1 h to 1 h 30 min",
    "1h30min a 2h": "1 h 30 min to 2 h",
    "2h a 2h30min": "2 h to 2 h 30 min",
    "2h30min a 3h": "2 h 30 min to 3 h",
    "Mais de 3h": "More than 3 h",
    "AEROLÍNEAS ARGENTINAS": "AEROLÍNEAS ARGENTINAS",
    "AEROMÉXICO": "AEROMÉXICO",
    "IBÉRIA": "IBÉRIA",
    "DEZEMBRO": "DECEMBER",
    "NOVEMBRO": "NOVEMBER",
    "OUTUBRO": "OCTOBER",
}


CODE_RE = re.compile(r"^[A-Z]{1,4}\d{0,5}[A-Z]?$|^[A-Z]\d{1,4}$")
ONLY_NUMBERISH_RE = re.compile(r"^[\d\s.,:/+%-]+$")
PORTUGUESE_HINT_RE = re.compile(
    r"[áàâãéêíóôõúüçÁÀÂÃÉÊÍÓÔÕÚÜÇ]|"
    r"\b(n[aã]o|sim|com|sem|para|pelo|pela|por|outros?|diretamente|"
    r"passageiro|aeroporto|embarque|desembarque|viagem|op[cç][oõ]es|"
    r"tempo|fila|qualidade|quantidade|facilidade|atendimento|"
    r"funcion[aá]rios|banheiros|bagagem|estacionamento|internet|"
    r"aplicativo|aplicativos|ensino|renda|familiar|anos|pessoas?|"
    r"valores|entre|mais|at[eé]|trabalho|lazer|dom[eé]stico|"
    r"formul[aá]rio|quest[aã]o|respostas|pesquisa|hora|data|"
    r"transporte|ve[ií]culo|carro|t[aá]xi|necess[aá]rio|qual|"
    r"senhor|senhora|faixa|grau|escolaridade|realizou|incluindo|"
    r"gostaria|registrar|algum|coment[aá]rio|adicional|servi[cç]os|"
    r"oferecidos|hor[aá]rio|chegou|conex[aã]o|conexao)\b",
    re.IGNORECASE,
)


def load_cache():
    if CACHE.exists():
        return json.loads(CACHE.read_text(encoding="utf-8"))
    return {}


def save_cache(cache):
    CACHE.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def normalize_translation(text):
    text = text.strip()
    text = re.sub(r"\bCia\. air\b", "airline", text, flags=re.IGNORECASE)
    text = re.sub(r"\bair company\b", "airline", text, flags=re.IGNORECASE)
    text = re.sub(r"\bmaís\b", "more", text, flags=re.IGNORECASE)
    text = re.sub(r"\bpiás\b", "sinks", text, flags=re.IGNORECASE)
    text = re.sub(r"\bSwrviço\b", "service", text, flags=re.IGNORECASE)
    text = text.replace(" x ", " x ")
    return text


def should_translate(value, sheet_name, row, col):
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or text.startswith("="):
        return False
    if row >= 4 and col in SKIP_COLUMNS.get(sheet_name, set()):
        return False
    if text in MANUAL:
        return True
    if ONLY_NUMBERISH_RE.match(text):
        return False
    if CODE_RE.match(text):
        return False
    return bool(PORTUGUESE_HINT_RE.search(text))


def translate_text(text, cache):
    if text in MANUAL:
        return MANUAL[text]
    if text in cache:
        return normalize_translation(cache[text])
    translated = argostranslate.translate.translate(text, "pt", "en")
    translated = normalize_translation(translated)
    cache[text] = translated
    return translated


def main():
    cache = load_cache()
    wb = load_workbook(SOURCE, data_only=False)

    candidates = []
    seen = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if should_translate(cell.value, ws.title, cell.row, cell.column):
                    if cell.value not in seen and cell.value not in cache:
                        seen.add(cell.value)
                        candidates.append(cell.value)

    print(f"unique strings to translate: {len(candidates)}")
    for i, text in enumerate(candidates, 1):
        translate_text(text, cache)
        if i % 100 == 0:
            save_cache(cache)
            print(f"translated {i}/{len(candidates)}")
    save_cache(cache)

    changed = 0
    for ws in wb.worksheets:
        original_title = ws.title
        for row in ws.iter_rows():
            for cell in row:
                if should_translate(cell.value, original_title, cell.row, cell.column):
                    new_value = translate_text(cell.value, cache)
                    if new_value != cell.value:
                        cell.value = new_value
                        changed += 1
        if original_title in MANUAL:
            ws.title = MANUAL[original_title]

    wb.save(TARGET)
    print(f"saved: {TARGET}")
    print(f"translated cells: {changed}")


if __name__ == "__main__":
    main()
