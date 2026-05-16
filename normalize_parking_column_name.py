from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
TARGET_HEADER = "UTILIZOU O ESTACIONAMENTO?"
NORMALIZED_HEADER = "UTILIZOU O ESTACIONAMENTO"

SEARCH_DIRS = [
    ROOT / "passenger_survey_downloads",
    ROOT / "passenger_survey_column_groups",
]


def normalize_workbook(path: Path) -> bool:
    wb = load_workbook(path)
    changed = False

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == TARGET_HEADER:
                    cell.value = NORMALIZED_HEADER
                    changed = True

    if changed:
        wb.save(path)

    return changed


def main() -> None:
    changed_files = []

    for directory in SEARCH_DIRS:
        if not directory.exists():
            continue
        for path in directory.rglob("*.xlsx"):
            if normalize_workbook(path):
                changed_files.append(path)

    for path in changed_files:
        print(path.relative_to(ROOT))

    print(f"changed files: {len(changed_files)}")


if __name__ == "__main__":
    main()
