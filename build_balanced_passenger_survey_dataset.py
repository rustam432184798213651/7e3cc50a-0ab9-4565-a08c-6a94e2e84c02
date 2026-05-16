import csv
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# csv используется не через pandas.read_csv, потому что исходные CSV отличаются
# разделителями и строкой заголовка. Модуль csv дает полный контроль над dialect.
# json нужен для сохранения отчета сборки и чтения кеша переводов.
# re нужен для нормализации пробелов и пустых строк через регулярные выражения.
# Path удобнее os.path: через оператор / собираются переносимые пути.
#
# numpy нужен для np.nan и проверки numpy-числовых типов.
# pandas является основной библиотекой табличной обработки.
# openpyxl читает XLSX напрямую, чтобы мы могли сами найти строку заголовков.
# Скрипт собирает единый ML-датасет из квартальных файлов бразильского
# passenger survey. На входе лежат файлы разных форматов и годов: часть CSV,
# часть XLSX, с немного отличающимися заголовками и большим числом пропусков,
# которые на самом деле означают "вопрос не задавался по маршруту анкеты".
# На выходе получается сбалансированный CSV с английскими названиями колонок,
# английскими категориальными значениями и бинарным target-признаком liked.
ROOT = Path(__file__).resolve().parent
# __file__ - путь к текущему .py файлу; resolve() делает путь абсолютным,
# parent берет директорию файла. Так скрипт можно запускать из любой cwd.
SOURCE_DIR = ROOT / "passenger_survey_downloads"
# Папка с исходными квартальными файлами. Оператор / у Path добавляет сегмент пути.
OUTPUT_DIR = ROOT / "balanced_passenger_survey_dataset"
# Папка, куда сохраняются итоговый CSV и JSON-отчет.

# Фиксированный random_state нужен, чтобы undersampling большинства класса
# и перемешивание строк давали воспроизводимый результат при каждом запуске.
RANDOM_STATE = 42
# Константа random_state фиксирует все случайные операции pandas sample().
TARGET_COLUMN = "liked"
# Имя бинарного target в итоговом датасете: 1 = удовлетворен, 0 = не удовлетворен.
SATISFACTION_COLUMN = "SATISFAÇÃO GERAL"
# Исходная колонка общей удовлетворенности, по ней строится TARGET_COLUMN.

# Эти поля не используются как признаки модели. Часть из них является
# идентификаторами или утечкой целевой переменной, часть слишком техническая
# для текущей постановки задачи, а airport/terminal специально удалены,
# чтобы модель не переобучалась на конкретные аэропорты/терминалы.
DROP_COLUMNS = [
    "CHAVE",
    "AEROPORTO",
    "TERMINAL",
    "ATENDIMENTO DA CIA. AÉREA2",
    SATISFACTION_COLUMN,
    "MOTIVO",
    "COMENTÁRIOS ADICIONAIS",
    "DATA",
    "INÍCIO COLETA",
    "FIM COLETA",
    "PORTÃO",
    "CIA AÉREA",
    "VOO",
    "source_file",
]

COLUMN_RENAMES = {
    "UTILIZOU O ESTACIONAMENTO?": "UTILIZOU O ESTACIONAMENTO",
}
# В разных кварталах один и тот же вопрос может иметь знак вопроса или не иметь.
# Ручное переименование приводит такие варианты к одному имени до объединения.

# Единое значение для структурной неприменимости вопроса. Это не обычный
# пропуск, а сигнал "этот вопрос не относился к данному пассажиру/процессу".
# Для числовых колонок ниже создается отдельный индикатор *_is_applicable.
NOT_APPLICABLE = "Not applicable"

# Эти значения мешают определить колонку как числовую, хотя содержательно
# они являются пропусками или ответами "не знаю/нет ответа".
NON_NUMERIC_MISSING_VALUES = {
    NOT_APPLICABLE,
    "NS",
    "NS/NR",
    "NS/NR/NA",
    "Unknown/no answer",
}

COLUMN_TRANSLATIONS = {
    "PROCESSO": "process",
    "AEROPORTO": "airport",
    "MÊS": "month",
    "TERMINAL": "terminal",
    "TIPO DE VOO": "flight_type",
    "CONEXÃO": "connection",
    "AQUISIÇÃO DA PASSAGEM": "ticket_purchased_by",
    "MEIO DE AQUISIÇÃO DA PASSAGEM": "ticket_purchase_channel",
    "MEIO DE  TRANSPORTE PARA O AEROPORTO": "transport_to_airport",
    "POSSUI DEFICIÊNCIA": "has_disability",
    "UTILIZA RECURSO ASSISTIVO": "uses_assistive_device",
    "SOLICITOU ASSISTÊNCIA ESPECIAL": "requested_special_assistance",
    "FORMA DE DESEMBARQUE UTILIZADA": "disembarkation_method_used",
    "AVALIAÇÃO DO MÉTODO DE DESEMBARQUE": "disembarkation_method_rating",
    "UTILIZOU O ESTACIONAMENTO": "used_parking",
    "FACILIDADE DE DESEMBARQUE NO MEIO-FIO": "curbside_dropoff_ease",
    "OPÇÕES DE TRANSPORTE ATÉ O AEROPORTO": "transport_options_to_airport",
    "FORMA DE CHECK-IN": "checkin_method",
    "PROCESSO DE CHECK IN": "checkin_process",
    "TEMPO DE ESPERA NA FILA": "checkin_queue_wait_time",
    "ORGANIZAÇÃO DAS FILAS": "checkin_queue_organization",
    "QUANTIDADE DE TOTENS AA": "self_service_kiosk_quantity",
    "QUANTIDADE DE BALCÕES": "checkin_counter_quantity",
    "CORDIALIDADE DOS FUNCIONÁRIOS": "staff_courtesy",
    "TEMPO DE ATENDIMENTO": "checkin_service_time",
    "PROCESSO DE AQUISIÇÃO DA PASSAGEM": "ticket_purchase_process",
    "ATENDIMENTO DA CIA. AÉREA": "airline_service",
    "PROCESSO DE INSPEÇÃO DE SEGURANÇA": "security_screening_process",
    "TEMPO DE ESPERA EM FILA": "security_queue_wait_time",
    "ORGANIZAÇÃO DAS FILAS2": "security_queue_organization",
    "ATENDIMENTO DOS FUNCIONÁRIOS": "security_staff_service",
    "CONTROLE MIGRATÓRIO": "immigration_control",
    "TEMPO DE ESPERA EM FILA3": "immigration_queue_wait_time",
    "ORGANIZAÇÃO DAS FILAS4": "immigration_queue_organization",
    "ATENDIMENTO DOS FUNCIONÁRIOS5": "immigration_staff_service",
    "QUANTIDADE DE GUICHÊS": "service_window_quantity",
    "CONTROLE ADUANEIRO": "customs_control",
    "TEMPO DE ESPERA EM FILA2": "customs_queue_wait_time",
    "ORGANIZAÇÃO DAS FILAS3": "customs_queue_organization",
    "ATENDIMENTO DOS FUNCIONÁRIOS4": "customs_staff_service",
    "ESTABELECIMENTOS DE ALIMENTAÇÃO": "food_beverage_outlets",
    "QUANTIDADE DE ESTABELECIMENTOS DE ALIMENTAÇÃO": "food_beverage_outlet_quantity",
    "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS DE ALIMENTAÇÃO": "food_beverage_quality_variety",
    "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS DE ALIMENTAÇÃO": "food_beverage_price_quality",
    "ESTABELECIMENTOS COMERCIAIS": "retail_outlets",
    "QUANTIDADE DE ESTABELECIMENTOS COMERCIAIS": "retail_outlet_quantity",
    "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS COMERCIAIS": "retail_quality_variety",
    "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS COMERCIAIS": "retail_price_quality",
    "ESTACIONAMENTO": "parking",
    "QUALIDADE DAS INSTALAÇÕES DE ESTACIONAMENTO": "parking_facility_quality",
    "FACILIDADE PARA ENCONTRAR VAGAS": "parking_space_availability_ease",
    "FACILIDADE DE ACESSO AO TERMINAL": "parking_terminal_access_ease",
    "RELAÇÃO CUSTO X BENEFÍCIO": "parking_value_for_money",
    "LOCALIZAÇÃO E DESLOCAMENTO": "location_and_movement",
    "SINALIZAÇÃO": "signage",
    "DISPONIBILIDADE DE PAINÉIS DE INFORMAÇÕES DE VOO": "flight_information_display_availability",
    "ACESSIBILIDADE DO TERMINAL": "terminal_accessibility",
    "CONFORTO DA SALA DE EMBARQUE": "boarding_lounge_comfort",
    "CONFORTO TÉRMICO": "thermal_comfort",
    "CONFORTO ACÚSTICO": "acoustic_comfort",
    "DISPONIBILIDADE DE ASSENTOS": "seat_availability",
    "DISPONIBILIDADE DE ASSENTOS RESERVADOS": "reserved_seat_availability",
    "DISPONIBILIDADE DE TOMADAS": "power_outlet_availability",
    "INTERNET DISPONIBILIZADA PELO AEROPORTO": "airport_internet",
    "VELOCIDADE DE CONEXÃO": "internet_connection_speed",
    "FACILIDADE DE ACESSO À REDE": "network_access_ease",
    "SANITÁRIOS": "restrooms",
    "QUANTIDADE DE BANHEIROS": "restroom_quantity",
    "LIMPEZA DOS BANHEIROS": "restroom_cleanliness",
    "MANUTENÇÃO GERAL DOS SANITÁRIOS": "restroom_maintenance",
    "LIMPEZA GERAL DO AEROPORTO": "overall_airport_cleanliness",
    "PROCESSO DE RESTITUIÇÃO DE BAGAGENS": "baggage_claim_process",
    "FACILIDADE DE IDENTIFICAÇÃO DA ESTEIRA DE RESTITUIÇÃO": "baggage_carousel_identification_ease",
    "TEMPO DE RESTITUIÇÃO": "baggage_claim_time",
    "INTEGRIDADE DA BAGAGEM": "baggage_integrity",
    "NACIONALIDADE": "nationality",
    "GÊNERO": "gender",
    "IDADE": "age_group",
    "ESCOLARIDADE": "education",
    "RENDA FAMILIAR": "household_income",
    "VIAJANDO SOZINHO": "traveling_alone",
    "NÚMERO DE ACOMPANHANTES": "number_of_companions",
    "MOTIVO DA VIAGEM": "trip_purpose",
    "QUANTIDADE DE VIAGENS NOS ÚLTIMOS 12 MESES": "trips_last_12_months",
    "JÁ EMBARCOU/DESEMBARCOU ANTES NO AEROPORTO": "used_airport_before_last_12_months",
    "ANTECEDÊNCIA": "arrival_lead_time",
    "TEMPO DE ESPERA": "connection_wait_time",
}
# Словарь переводит исходные португальские названия колонок в короткие
# английские snake_case имена. Ключи - имена в сырых файлах, значения -
# имена в итоговом CSV. Если колонки здесь нет, ее имя останется без изменений.


def normalize_text(value) -> str:
    # Приводит произвольное значение к нижнему регистру и нормализует пробелы.
    # Эта функция используется почти всеми rule-based нормализаторами ниже,
    # чтобы одинаково обрабатывать регистр, лишние пробелы и NaN.
    if pd.isna(value):
        # pd.isna ловит np.nan, None и pandas NA. Для текстовой нормализации
        # такие значения превращаем в пустую строку, чтобы дальше не падать.
        return ""
    text = str(value).strip().lower()
    # str(value) делает значение строкой; strip() убирает пробелы по краям;
    # lower() приводит к нижнему регистру, чтобы сравнения были case-insensitive.
    text = re.sub(r"\s+", " ", text)
    # \s+ означает "один или больше любых whitespace-символов".
    # Заменяем подряд идущие пробелы/табуляции/переносы строк одним пробелом.
    return text


def normalize_trip_purpose(value) -> str:
    # В исходных данных цель поездки часто заполнена свободным текстом.
    # Здесь такие ответы группируются в ограниченный набор категорий, чтобы
    # one-hot encoding не создавал сотни редких и шумных признаков.
    text = normalize_text(value)
    if not text:
        return np.nan

    if text in {"lazer", "turismo", "passeio", "férias", "ferias", "descanso", "carnaval", "show", "lua de mel"}:
        return "Leisure"
    if "lazer" in text and "trabalho" in text:
        return "Leisure and business"
    if any(token in text for token in ["trabalho", "negocio", "negócio", "treinamento", "convenção"]):
        return "Business"
    if any(token in text for token in ["familiar", "família", "familia", "parentes", "filho", "filha", "mãe", "mae", "pais"]):
        return "Family"
    if any(token in text for token in ["estudo", "congresso", "seminário", "seminario", "palestra", "curso", "faculdade", "tcc", "pós", "pos "]):
        return "Study/conference"
    if any(token in text for token in ["saúde", "saude", "tratamento", "consulta", "exame", "médic", "medic", "cirurgia", "doença"]):
        return "Health"
    if any(token in text for token in ["funeral", "óbito", "obito", "velório", "velorio", "falecimento", "enterro", "morte"]):
        return "Funeral"
    if any(token in text for token in ["mudança", "mudanca", "morar", "residência", "residencia", "domicílio", "domicilio"]):
        return "Relocation/return home"
    if any(token in text for token in ["retorno", "voltando", "volta", "regresso", "casa"]):
        return "Relocation/return home"
    if any(token in text for token in ["esporte", "campeonato", "competição", "competicao", "atleta", "sport"]):
        return "Sports"
    if any(token in text for token in ["jurídico", "juridico", "embaixada", "consulado", "visto", "passaporte", "green card", "document"]):
        return "Legal/embassy"
    if any(token in text for token in ["casamento", "noivado"]):
        return "Wedding"
    if any(token in text for token in ["relig", "igreja", "missão", "missao", "promessa"]):
        return "Religious/mission"
    if "concurso" in text or "processo seletivo" in text or "perícia" in text or "pericia" in text:
        return "Exam/selection process"
    if "compras" in text or "compra" in text:
        return "Shopping"
    if "ns/nr" in text or "não quis informar" in text or "nao quis informar" in text:
        return "Unknown/no answer"
    if text.startswith("outros"):
        return "Other"
    return "Other"


def normalize_transport_to_airport(value) -> str:
    # Аналогично цели поездки, транспорт до аэропорта содержит много вариантов
    # свободного ввода: Uber, такси, личный автомобиль, трансфер и т.д.
    # Нормализуем их в устойчивые категории.
    text = normalize_text(value)
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE

    if any(token in text for token in ["aplicativo", "uber", "99", "ub r"]):
        return "Ride-hailing app"
    if any(token in text for token in ["carro próprio", "carro proprio", "carro particular", "veículo particular", "veiculo particular"]):
        return "Own car"
    if any(token in text for token in ["carona", "someone picked"]):
        return "Ride from someone"
    if "táxi" in text or "taxi" in text:
        return "Taxi"
    if any(token in text for token in ["ônibus", "onibus", "bus"]):
        return "Bus"
    if any(token in text for token in ["metrô", "metro", "trem", "vlt", "aeromóvel", "aeromovel"]):
        return "Public rail/transit"
    if any(token in text for token in ["veículo alugado", "veiculo alugado", "carro alugado"]):
        return "Rental car"
    if "transfer" in text or "transffer" in text or "translado" in text:
        return "Transfer/shuttle"
    if "van" in text or "vam" in text:
        return "Van/shuttle"
    if any(token in text for token in ["empresa", "trabalho", "prefeitura", "exército", "exercito"]):
        return "Company/official vehicle"
    if any(token in text for token in ["moto", "motocicleta", "mototaxi", "moto táxi"]):
        return "Motorcycle"
    if any(token in text for token in ["avião", "aviao", "plane"]):
        return "Airplane"
    if any(token in text for token in ["andando", "a pé", "a pe"]):
        return "Walking"
    if "helicóptero" in text or "helicoptero" in text:
        return "Helicopter"
    if "ns/nr" in text or "não sabe" in text or "nao sabe" in text:
        return "Unknown/no answer"
    if text.startswith("outros") or text in {"outros", "transporte alternativo"}:
        return "Other"
    return "Other"


def normalize_simple_category(value, mapping, default=None):
    # Универсальная функция для колонок, где достаточно словаря соответствий:
    # португальское значение -> английское значение. Если значения нет
    # в mapping, возвращается исходная строка или заданный default.
    if pd.isna(value):
        return np.nan
    original = str(value).strip()
    text = normalize_text(original).rstrip(".")
    # rstrip(".") убирает точку только справа. В исходных файлах одно и то же
    # значение может встречаться как "Não" и "Não.", это не должны быть
    # разные категории.
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    return mapping.get(text, default if default is not None else original)
    # mapping.get ищет нормализованное значение в словаре. Если ключа нет,
    # возвращаем default, если он задан, иначе исходное значение без пробелов.


def normalize_ticket_purchase_channel(value) -> str:
    # Канал покупки билета тоже содержит свободные уточнения. Нормализуем
    # конкретные варианты в более широкие группы: сайт/приложение авиакомпании,
    # агентство, телефон, физический офис, мили и т.д.
    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    if "internet" in text or "aplicativo" in text:
        return "Airline website/app"
    if "agência" in text or "agencia" in text or "viajanet" in text or "decolar" in text or "skay" in text:
        return "Travel agency/third-party website"
    if "telefone" in text:
        return "Airline phone"
    if "loja física" in text or "loja fisica" in text:
        return "Airline physical store"
    if "milhas" in text or "smiles" in text or "smails" in text:
        return "Miles/loyalty program"
    if "banco" in text or "cartao" in text or "cartão" in text:
        return "Bank/credit card benefit"
    if "funcionario" in text or "funcionário" in text or "tripulação" in text:
        return "Employee/crew benefit"
    if "ns" in text or "não sabe" in text or "nao lembra" in text:
        return "Unknown/no answer"
    if "whatsapp" in text:
        return "Messaging app"
    if "stand by" in text:
        return "Standby"
    return "Other"


def normalize_assistive_device(value) -> str:
    # Сводит разные формулировки ассистивных средств к единому набору
    # категорий, важному для анализа пассажиров с инвалидностью.
    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    if text in {"nenhum", "nenhuma"}:
        return "None"
    if "cadeira de rodas motorizada" in text:
        return "Motorized wheelchair"
    if "cadeira de rodas" in text:
        return "Manual wheelchair"
    if "bengala" in text:
        return "Cane"
    if "muleta" in text:
        return "Crutches"
    if "andador" in text:
        return "Walker"
    if "cão-guia" in text or "cao-guia" in text or "guia" in text:
        return "Guide dog"
    if "aparelho auditivo" in text:
        return "Hearing aid"
    if "scooter" in text:
        return "Scooter"
    if "medic" in text or "remedio" in text or "remédio" in text:
        return "Medication"
    if "protese" in text or "prótese" in text or "palmilha" in text or "bota" in text:
        return "Prosthesis/orthopedic aid"
    if "acompanhante" in text or "auxilio" in text or "auxílio" in text:
        return "Personal assistance"
    return "Other"


def normalize_special_assistance(value) -> str:
    # Нормализует ответы о специальной помощи: была ли запрошена, за сколько
    # дней, либо почему пассажир ее не запросил.
    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    if "sim" in text and "menos de 2" in text:
        return "Yes: less than 2 days in advance"
    if "sim" in text and "entre 2 e 3" in text:
        return "Yes: 2 to 3 days in advance"
    if "sim" in text and "mais de 3" in text:
        return "Yes: more than 3 days in advance"
    if "sim" in text:
        return "Yes"
    if "não quis" in text or "nao quis" in text:
        return "Did not want"
    if "não sabia" in text or "nao sabia" in text or "não lembro" in text:
        return "Did not know/remember"
    if "não conseguiu" in text or "nao conseguiu" in text:
        return "Could not request"
    if "não precisa" in text or "nao precisa" in text or "necessidade" in text or "não preciso" in text:
        return "Did not need"
    if "ns" in text:
        return "Unknown/no answer"
    return "Other"


def normalize_disembarkation_method(value) -> str:
    # Приводит способ высадки к компактным категориям: телетрап, мобильная
    # рампа, удаленная стоянка, кресло-коляска, лестница и т.д.
    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    if "ponte" in text:
        return "Jet bridge"
    if "rampa" in text:
        return "Mobile ramp"
    if "remoto" in text or "ônibus" in text or "onibus" in text:
        return "Remote by bus/walking"
    if "cadeira robótica" in text or "cadeira robotica" in text:
        return "Robotic chair"
    if "cadeira de rodas" in text:
        return "Wheelchair"
    if "escada" in text:
        return "Stairs"
    if "ambulift" in text:
        return "Ambulift"
    if "ns" in text:
        return "Unknown/no answer"
    if "normal" in text or "nenhum" in text:
        return "None/standard"
    return "Other"


def normalize_checkin_method(value) -> str:
    # Объединяет варианты check-in, включая смешанные варианты вроде
    # "internet/app + counter" или "self-service kiosk + counter".
    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        return NOT_APPLICABLE
    if "internet" in text and "balc" in text:
        return "Internet/app + counter"
    if "totem" in text and "balc" in text:
        return "Self-service kiosk + counter"
    if "internet" in text or "aplicativo" in text:
        return "Internet/app"
    if "balc" in text:
        return "Counter"
    if "totem" in text:
        return "Self-service kiosk"
    return "Other"


def normalize_household_income(value) -> str:
    # Доход хранится как диапазоны минимальных зарплат. Здесь приводим
    # разные написания диапазонов к единому английскому виду.
    text = normalize_text(value)
    if not text:
        return np.nan
    if "prefiro" in text:
        return "Prefer not to say"
    if "até 1" in text or "ate 1" in text:
        return "Up to 1 minimum wage"
    if "entre 1 e 2" in text:
        return "Between 1 and 2 minimum wages"
    if "entre 2 e 4" in text:
        return "Between 2 and 4 minimum wages"
    if "entre 4 e 10" in text:
        return "Between 4 and 10 minimum wages"
    if "entre 10 e 20" in text:
        return "Between 10 and 20 minimum wages"
    if "mais de 20" in text or "more than 20" in text:
        return "More than 20 minimum wages"
    return "Other"


def normalize_time_range_to_seconds(value):
    # В анкете время прихода заранее и ожидание пересадки записаны диапазонами.
    # Для моделей удобнее числовой признак, поэтому диапазоны переводятся
    # в секунды. Для закрытых интервалов берется середина, а "More than 3 h"
    # кодируется нижней границей 3 часа = 10800 секунд.
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        # Если значение уже числовое, не трогаем его. Это делает функцию
        # устойчивой к повторному запуску или смешанным типам в колонке.
        return value

    text = normalize_text(value).rstrip(".")
    if not text:
        return np.nan
    if text == normalize_text(NOT_APPLICABLE):
        # Для Not applicable возвращаем строковый маркер, а не NaN. Позже
        # impute_remaining_missing создаст отдельный *_is_applicable индикатор.
        return NOT_APPLICABLE
    if text in {"ns", "ns/nr", "unknown/no answer", "conforme necessário", "as applicable"}:
        return np.nan

    mapping = {
        "30min a 1h": 45 * 60,
        "30 min to 1 h": 45 * 60,
        "1h a 1h30min": 75 * 60,
        "1 h to 1 h 30 min": 75 * 60,
        "1h30min a 2h": 105 * 60,
        "1 h 30 min to 2 h": 105 * 60,
        "2h a 2h30min": 135 * 60,
        "2 h to 2 h 30 min": 135 * 60,
        "2h30min a 3h": 165 * 60,
        "2 h 30 min to 3 h": 165 * 60,
        "mais de 3h": 180 * 60,
        "more than 3 h": 180 * 60,
    }
    return mapping.get(text, np.nan)
    # Если встретился неизвестный формат времени, превращаем его в NaN:
    # это обычный пропуск, который будет заполнен медианой.


def normalize_final_categorical_values(df: pd.DataFrame) -> pd.DataFrame:
    # Финальный проход по уже переименованным колонкам. На этом этапе
    # часть значений уже нормализована вручную, но остаются простые
    # категориальные поля вроде месяца, пола, возраста и национальности.
    df = df.copy()
    simple_mappings = {
        "process": {"embarque": "Boarding", "desembarque": "Disembarkation"},
        "month": {
            "janeiro": "January",
            "fevereiro": "February",
            "março": "March",
            "marco": "March",
            "abril": "April",
            "maio": "May",
            "junho": "June",
            "julho": "July",
            "agosto": "August",
            "setembro": "September",
            "outubro": "October",
            "novembro": "November",
            "dezembro": "December",
        },
        "flight_type": {"doméstico": "Domestic", "domestico": "Domestic", "internacional": "International"},
        "connection": {"em conexão": "Connecting", "em conexao": "Connecting", "embarcado no aeroporto": "Boarded at the airport"},
        "ticket_purchased_by": {"pelo passageiro": "By the passenger", "por terceiros": "By third parties"},
        "has_disability": {
            "nenhuma": "No",
            "deficiência física / motora": "Physical/motor disability",
            "deficiência física/motora": "Physical/motor disability",
            "deficiência física /motora": "Physical/motor disability",
            "deficiência visual": "Visual disability",
            "deficiência auditiva / surdez": "Hearing disability/deafness",
            "deficiência auditiva/surdez": "Hearing disability/deafness",
            "deficiência mental/intelectual": "Mental/intellectual disability",
            "deficiência mental / intelectual": "Mental/intellectual disability",
        },
        "used_parking": {"sim": "Yes", "não": "No", "nao": "No"},
        "nationality": {"brasileira": "Brazilian", "brasileiro": "Brazilian", "estrangeira": "Foreign", "estrangeiro": "Foreign"},
        "gender": {"masculino": "Male", "feminino": "Female"},
        "age_group": {
            "18 a 25 anos": "18 to 25 years",
            "26 a 35 anos": "26 to 35 years",
            "36 a 45 anos": "36 to 45 years",
            "46 a 55 anos": "46 to 55 years",
            "56 a 64 anos": "56 to 65 years",
            "56 a 65 anos": "56 to 65 years",
            "mais de 65 anos": "Over 65 years",
        },
        "education": {
            "analfabeto": "No formal literacy",
            "ensino fundamental": "Elementary education",
            "ensino médio": "High school",
            "superior": "Higher education",
            "especialização de nível superior": "Graduate specialization",
            "mestrado": "Master's degree",
            "doutorado": "Doctorate",
            "master's degree": "Master's degree",
        },
        "traveling_alone": {"sim": "Yes", "não": "No", "nao": "No"},
        "number_of_companions": {
            "1 pessoa": "1 person",
            "2 pessoas": "2 people",
            "3 pessoas": "3 people",
            "4 ou mais pessoas": "4 or more people",
        },
        "trips_last_12_months": {
            "1 - primeira viagem": "1 - first trip",
            "1 (primeira viagem)": "1 - first trip",
            "2 a 3": "2 to 3",
            "4 a 5": "4 to 5",
            "6 a 10": "6 to 10",
            "mais de 11": "More than 11",
            "ns/nr": "Unknown/no answer",
        },
        "used_airport_before_last_12_months": {"sim": "Yes", "não": "No", "nao": "No"},
    }
    for col, mapping in simple_mappings.items():
        # items() возвращает пары "имя колонки -> словарь переводов".
        if col in df.columns:
            df[col] = df[col].map(lambda value: normalize_simple_category(value, mapping))
            # Series.map применяет функцию к каждому значению колонки.

    special_normalizers = {
        "ticket_purchase_channel": normalize_ticket_purchase_channel,
        "uses_assistive_device": normalize_assistive_device,
        "requested_special_assistance": normalize_special_assistance,
        "disembarkation_method_used": normalize_disembarkation_method,
        "checkin_method": normalize_checkin_method,
        "household_income": normalize_household_income,
    }
    for col, normalizer in special_normalizers.items():
        # special_normalizers хранит функции для колонок, где простого словаря
        # мало и нужна rule-based логика по подстрокам.
        if col in df.columns:
            df[col] = df[col].map(normalizer)

    rating_like_categorical = [
        "reserved_seat_availability",
        "power_outlet_availability",
    ]
    for col in rating_like_categorical:
        if col in df.columns:
            df[col] = df[col].replace({"NS": "NS/NR", "NS/NR/NA": "NS/NR"})

    return df


def load_translation_cache() -> dict:
    # Кеш создается отдельным скриптом перевода Excel. Здесь он используется
    # только как локальный словарь точных переводов, без сетевых запросов
    # и без повторного машинного перевода.
    cache_path = ROOT / "translation_cache_pt_en.json"
    # Путь к JSON-файлу с уже готовыми переводами pt->en.
    if not cache_path.exists():
        # Если кеша нет, возвращаем пустой словарь и просто пропускаем
        # кешированный перевод. Скрипт остается работоспособным.
        return {}
    # read_text читает файл как строку. encoding="utf-8" нужен для португальских
    # символов. json.loads превращает JSON-строку в Python dict.
    return json.loads(cache_path.read_text(encoding="utf-8"))


def normalize_cached_translation(value):
    # Если строковое значение осталось на португальском и есть в кеше,
    # заменяем его английским вариантом. Несколько частых коротких ответов
    # переведены вручную, потому что они встречаются в разных регистрах
    # и не всегда должны идти через общий кеш.
    if not isinstance(value, str):
        return value

    text = value.strip()
    # Здесь не используем lower(), потому что ключи кеша чувствительны к точному
    # исходному тексту. Сначала пробуем точное совпадение.
    if not text:
        return value

    manual_translations = {
        "Outro": "Other",
        "Outros": "Other",
        "Não": "No",
        "Nao": "No",
        "Sim": "Yes",
        "NS": "NS/NR",
        "NS/NR": "Unknown/no answer",
    }
    if text in manual_translations:
        return manual_translations[text]

    cache = normalize_cached_translation.cache
    # cache хранится как атрибут функции, чтобы не передавать словарь через
    # каждый вызов Series.map.
    if text in cache:
        return cache[text]
    return value


normalize_cached_translation.cache = {}
# Инициализируем атрибут функции пустым словарем. Позже
# translate_remaining_text_values заменит его реальным кешем.


def translate_remaining_text_values(df: pd.DataFrame) -> pd.DataFrame:
    # Проходит по всем текстовым колонкам и добивает оставшиеся точные
    # португальские значения через translation_cache_pt_en.json.
    df = df.copy()
    # copy защищает входной DataFrame от побочных изменений.
    normalize_cached_translation.cache = load_translation_cache()
    if not normalize_cached_translation.cache:
        return df

    for col in df.select_dtypes(include=["object", "str"]).columns:
        # select_dtypes выбирает только текстовые колонки. Числовые признаки
        # переводить нельзя, иначе можно случайно превратить числа в строки.
        df[col] = df[col].map(normalize_cached_translation)
    return df


def normalize_header_cell(value) -> str:
    # Заголовки в исходных файлах могут быть None или содержать пробелы.
    # Приводим их к строкам, чтобы дальнейшая нормализация была стабильной.
    if value is None:
        return ""
    return str(value).strip()


def normalize_columns(columns):
    # Делает имена колонок пригодными для pandas: убирает пустые заголовки,
    # применяет ручные переименования и добавляет суффиксы .1, .2 для дублей.
    # В анкетах есть повторяющиеся названия вроде "TEMPO DE ESPERA EM FILA",
    # относящиеся к разным блокам анкеты.
    normalized = []
    seen = {}
    for col in columns:
        name = normalize_header_cell(col)
        # normalize_header_cell превращает None в "" и убирает пробелы.
        name = COLUMN_RENAMES.get(name, name)
        # Применяем ручное переименование, если такой заголовок есть в словаре.
        if not name:
            name = "Unnamed"
            # Пустой заголовок нельзя использовать как нормальное имя колонки.
        count = seen.get(name, 0)
        # seen хранит, сколько раз такое имя уже встречалось.
        seen[name] = count + 1
        if count:
            name = f"{name}.{count}"
            # pandas сам делает похожие суффиксы для дублей. Здесь мы делаем
            # это явно, потому что DataFrame собирается вручную из rows.
        normalized.append(name)
    return normalized


def decode_csv_bytes(data: bytes) -> str:
    # Старые CSV могут быть сохранены в разных кодировках. Сначала пробуем
    # UTF-8 с BOM, затем обычный UTF-8, затем latin1 как наиболее безопасный
    # fallback для португальских файлов.
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            # utf-8-sig дополнительно удаляет BOM, если файл сохранен Excel.
            return data.decode(encoding)
        except UnicodeDecodeError:
            # Если кодировка не подошла, пробуем следующую.
            continue
    # errors="replace" заменяет нечитаемые байты спецсимволом вместо падения.
    return data.decode("latin1", errors="replace")


def find_header_index(rows):
    # В разных файлах строка заголовков находится не всегда на первой строке.
    # Ищем явные маркеры анкеты ("PROCESSO" или "CHAVE"), а если их нет,
    # берем самую заполненную строку среди первых 10 строк.
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        # Проверяем только первые 10 строк: заголовки находятся в начале файла,
        # а просмотр всего файла был бы лишним.
        values = [normalize_header_cell(value) for value in row]
        score = sum(bool(value) for value in values)
        # score - число непустых ячеек в строке. Если явных маркеров нет,
        # самая заполненная строка, скорее всего, является заголовком.
        if "PROCESSO" in values or "CHAVE" in values:
            return index
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def read_csv_resource(path: Path) -> pd.DataFrame:
    # CSV-файлы отличаются разделителями, поэтому используем csv.Sniffer.
    # После определения dialect вручную собираем DataFrame с найденной
    # строкой заголовка.
    text = decode_csv_bytes(path.read_bytes())
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
        # sniff анализирует первые 8192 символа и пытается определить разделитель.
        # delimiters ограничивает возможные варианты: запятая, точка с запятой,
        # табуляция или вертикальная черта.
    except csv.Error:
        dialect = csv.excel
        # Если определить разделитель не удалось, используем стандарт CSV Excel.

    rows = list(csv.reader(text.splitlines(), dialect))
    # splitlines сохраняет строки без символов переноса; csv.reader разбирает
    # кавычки и разделители согласно найденному dialect.
    header_index = find_header_index(rows)
    columns = normalize_columns(rows[header_index])
    data_rows = rows[header_index + 1 :]
    # Все строки после header_index считаются данными.
    df = pd.DataFrame(data_rows, columns=columns)
    # columns задает имена колонок DataFrame; data_rows - матрица значений.
    return df


def read_xlsx_resource(path: Path) -> pd.DataFrame:
    # В Excel-файлах лист LEGENDA содержит описание анкеты, а данные лежат
    # на первом рабочем листе. Читаем значения через openpyxl, чтобы одинаково
    # обработать XLSX и сохранить контроль над строкой заголовков.
    wb = load_workbook(path, read_only=True, data_only=True)
    # read_only=True экономит память при чтении больших XLSX.
    # data_only=True возвращает значения ячеек, а не формулы.
    sheet_names = [name for name in wb.sheetnames if name.upper() != "LEGENDA"]
    # Лист LEGENDA содержит справочник/описание, его не надо включать в данные.
    ws = wb[sheet_names[0]]
    # Берем первый не-LEGENDA лист: в этих файлах именно он содержит ответы.
    rows = list(ws.iter_rows(values_only=True))
    # values_only=True возвращает простые Python-значения вместо объектов Cell.
    header_index = find_header_index(rows)
    columns = normalize_columns(rows[header_index])
    data_rows = rows[header_index + 1 :]
    df = pd.DataFrame(data_rows, columns=columns)
    return df


def read_resource(path: Path) -> pd.DataFrame:
    # Единая точка чтения файла: выбирает парсер по расширению.
    if path.suffix.lower() == ".csv":
        return read_csv_resource(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_xlsx_resource(path)
    raise ValueError(f"Unsupported file type: {path}")


def clean_empty_values(df: pd.DataFrame) -> pd.DataFrame:
    # Удаляет полностью пустые строки, превращает пустые строки в NaN
    # и обрезает пробелы в текстовых ячейках.
    df = df.copy()
    df = df.replace(r"^\s*$", np.nan, regex=True)
    # regex=True включает регулярное выражение. ^\s*$ означает строку, состоящую
    # только из пробелов или вообще пустую. Такие значения превращаются в NaN.
    df = df.dropna(axis=0, how="all")
    # axis=0 означает "удалять строки"; how="all" - только если вся строка NaN.
    for col in df.select_dtypes(include=["object", "str"]).columns:
        df[col] = df[col].map(
            lambda value: value.strip() if isinstance(value, str) else value
        )
        # strip применяется только к строкам; NaN и числа оставляются как есть.
    return df


def score_to_numeric(series: pd.Series) -> pd.Series:
    # Рейтинги 1-5 иногда содержат NS/NR. Для числовой логики такие ответы
    # должны стать NaN, а не отдельной строковой категорией.
    # errors="coerce" означает: все, что нельзя преобразовать в число,
    # заменить на NaN, а не выбрасывать исключение.
    return pd.to_numeric(series.replace({"NS/NR": np.nan, "NS": np.nan}), errors="coerce")


def is_rating_leq_3(df: pd.DataFrame, col: str) -> pd.Series:
    # По методологии SAC дочерние вопросы задаются только если оценка
    # родительского макроиндикатора равна 1, 2 или 3.
    if col not in df.columns:
        return pd.Series(False, index=df.index)
        # Возвращаем Series той же длины, чтобы ее можно было использовать
        # как маску для loc. False означает "не помечать ничего".
    score = score_to_numeric(df[col])
    return score.isin([1, 2, 3])
    # isin([1, 2, 3]) возвращает boolean Series: True для оценок 1, 2 или 3.


def is_yes(series: pd.Series) -> pd.Series:
    # Утилита для португальского "sim" в routing-логике.
    return series.astype(str).str.strip().str.lower().eq("sim")
    # astype(str) нужен, чтобы .str методы работали даже при NaN/числах.
    # eq("sim") - векторное сравнение каждого значения с португальским "да".


def has_disability(df: pd.DataFrame) -> pd.Series:
    # Возвращает маску пассажиров, для которых блок вопросов об ассистивных
    # средствах и специальной помощи действительно применим.
    if "POSSUI DEFICIÊNCIA" not in df.columns:
        return pd.Series(False, index=df.index)
    value = df["POSSUI DEFICIÊNCIA"].astype(str).str.strip().str.lower()
    return value.notna() & ~value.isin(["nenhuma", "nan", "none", ""])
    # ~ инвертирует boolean Series. Итоговая маска True там, где значение
    # не равно "нет инвалидности" и не является пустым/служебным.


def mark_not_applicable(df: pd.DataFrame, columns, applicable_mask) -> None:
    # Проставляет NOT_APPLICABLE там, где вопрос по логике анкеты не должен
    # был задаваться. Это лучше, чем обычный NaN: модель получает отдельный
    # сигнал применимости через *_is_applicable для числовых колонок.
    for col in columns:
        if col in df.columns:
            df.loc[~applicable_mask, col] = NOT_APPLICABLE
            # loc[mask, col] меняет только строки, где mask=True.
            # Здесь используется ~applicable_mask: помечаем НЕприменимые строки.


def apply_questionnaire_routing(df: pd.DataFrame) -> pd.DataFrame:
    # В исходной анкете не каждый вопрос задается каждому пассажиру.
    # Например, вопросы посадки не относятся к высадке, парковка не относится
    # к тем, кто ее не использовал, а дочерние вопросы блока задаются только
    # при определенной оценке родительского вопроса. Этот блок восстанавливает
    # такую маршрутизацию и явно кодирует неприменимость вопроса.
    df = df.copy()

    process = df["PROCESSO"].astype(str).str.strip().str.lower()
    # PROCESSO в исходной анкете определяет тип процесса: embarque или desembarque.
    is_boarding = process.eq("embarque")
    is_disembarkation = process.eq("desembarque")
    # eq возвращает boolean Series. Эти маски дальше управляют routing-логикой.

    boarding_only = [
        "PORTÃO",
        "VOO",
        "CONEXÃO",
        "AQUISIÇÃO DA PASSAGEM",
        "MEIO DE AQUISIÇÃO DA PASSAGEM",
        "MEIO DE  TRANSPORTE PARA O AEROPORTO",
        "UTILIZOU O ESTACIONAMENTO",
        "FACILIDADE DE DESEMBARQUE NO MEIO-FIO",
        "OPÇÕES DE TRANSPORTE ATÉ O AEROPORTO",
        "FORMA DE CHECK-IN",
        "PROCESSO DE CHECK IN",
        "PROCESSO DE AQUISIÇÃO DA PASSAGEM",
        "PROCESSO DE INSPEÇÃO DE SEGURANÇA",
        "ESTABELECIMENTOS DE ALIMENTAÇÃO",
        "ESTABELECIMENTOS COMERCIAIS",
        "ESTACIONAMENTO",
        "CONFORTO DA SALA DE EMBARQUE",
        "DISPONIBILIDADE DE PAINÉIS DE INFORMAÇÕES DE VOO",
        "DISPONIBILIDADE DE TOMADAS",
        "INTERNET DISPONIBILIZADA PELO AEROPORTO",
        "SANITÁRIOS",
        "JÁ EMBARCOU/DESEMBARCOU ANTES NO AEROPORTO",
        "ANTECEDÊNCIA",
        "TEMPO DE ESPERA",
    ]
    mark_not_applicable(df, boarding_only, is_boarding)
    # boarding_only применим только к посадке. Для строк высадки эти вопросы
    # получают Not applicable.

    disembarkation_only = [
        "FORMA DE DESEMBARQUE UTILIZADA",
        "AVALIAÇÃO DO MÉTODO DE DESEMBARQUE",
        "CONTROLE ADUANEIRO",
        "PROCESSO DE RESTITUIÇÃO DE BAGAGENS",
        "FACILIDADE DE IDENTIFICAÇÃO DA ESTEIRA DE RESTITUIÇÃO",
        "TEMPO DE RESTITUIÇÃO",
        "INTEGRIDADE DA BAGAGEM",
    ]
    mark_not_applicable(df, disembarkation_only, is_disembarkation)
    # disembarkation_only применим только к высадке. Для строк посадки эти
    # вопросы получают Not applicable.

    if "CONEXÃO" in df.columns:
        connection = df["CONEXÃO"].astype(str).str.strip().str.lower()
        boarded_at_airport = is_boarding & connection.eq("embarcado no aeroporto")
        connecting = is_boarding & connection.eq("em conexão")
        # & делает поэлементное логическое "и": условие должно быть истинно
        # и для типа процесса, и для конкретного значения connection.
        mark_not_applicable(
            df,
            [
                "MEIO DE  TRANSPORTE PARA O AEROPORTO",
                "ANTECEDÊNCIA",
            ],
            boarded_at_airport,
        )
        mark_not_applicable(df, ["TEMPO DE ESPERA"], connecting)

    disability = has_disability(df)
    # disability=True только для пассажиров, у которых указана инвалидность.
    mark_not_applicable(
        df,
        [
            "UTILIZA RECURSO ASSISTIVO",
            "SOLICITOU ASSISTÊNCIA ESPECIAL",
            "AVALIAÇÃO DO MÉTODO DE DESEMBARQUE",
            "ACESSIBILIDADE DO TERMINAL",
            "DISPONIBILIDADE DE ASSENTOS RESERVADOS",
            "ATENDIMENTO DA CIA. AÉREA",
        ],
        disability,
    )

    if "UTILIZOU O ESTACIONAMENTO" in df.columns:
        mark_not_applicable(
            df,
            [
                "ESTACIONAMENTO",
                "QUALIDADE DAS INSTALAÇÕES DE ESTACIONAMENTO",
                "FACILIDADE PARA ENCONTRAR VAGAS",
                "FACILIDADE DE ACESSO AO TERMINAL",
                "RELAÇÃO CUSTO X BENEFÍCIO",
            ],
            is_boarding & is_yes(df["UTILIZOU O ESTACIONAMENTO"]),
            # Парковочные вопросы применимы только для посадки и только если
            # пассажир ответил "Sim" на вопрос об использовании парковки.
        )

    parent_to_children = {
        "PROCESSO DE CHECK IN": [
            "TEMPO DE ESPERA NA FILA",
            "ORGANIZAÇÃO DAS FILAS",
            "QUANTIDADE DE TOTENS AA",
            "QUANTIDADE DE BALCÕES",
            "CORDIALIDADE DOS FUNCIONÁRIOS",
            "TEMPO DE ATENDIMENTO",
        ],
        "PROCESSO DE INSPEÇÃO DE SEGURANÇA": [
            "TEMPO DE ESPERA EM FILA",
            "ORGANIZAÇÃO DAS FILAS2",
            "ATENDIMENTO DOS FUNCIONÁRIOS",
        ],
        "CONTROLE MIGRATÓRIO": [
            "TEMPO DE ESPERA EM FILA3",
            "ORGANIZAÇÃO DAS FILAS4",
            "ATENDIMENTO DOS FUNCIONÁRIOS5",
            "QUANTIDADE DE GUICHÊS",
        ],
        "CONTROLE ADUANEIRO": [
            "TEMPO DE ESPERA EM FILA2",
            "ORGANIZAÇÃO DAS FILAS3",
            "ATENDIMENTO DOS FUNCIONÁRIOS4",
        ],
        "ESTABELECIMENTOS DE ALIMENTAÇÃO": [
            "QUANTIDADE DE ESTABELECIMENTOS DE ALIMENTAÇÃO",
            "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS DE ALIMENTAÇÃO",
            "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS DE ALIMENTAÇÃO",
        ],
        "ESTABELECIMENTOS COMERCIAIS": [
            "QUANTIDADE DE ESTABELECIMENTOS COMERCIAIS",
            "QUALIDADE E VARIEDADE DE OPÇÕES DE ESTABELECIMENTOS COMERCIAIS",
            "RELAÇÃO PREÇO x QUALIDADE DOS ESTABELECIMENTOS COMERCIAIS",
        ],
        "LOCALIZAÇÃO E DESLOCAMENTO": [
            "SINALIZAÇÃO",
            "DISPONIBILIDADE DE PAINÉIS DE INFORMAÇÕES DE VOO",
            "ACESSIBILIDADE DO TERMINAL",
        ],
        "CONFORTO DA SALA DE EMBARQUE": [
            "CONFORTO TÉRMICO",
            "CONFORTO ACÚSTICO",
            "DISPONIBILIDADE DE ASSENTOS",
        ],
        "INTERNET DISPONIBILIZADA PELO AEROPORTO": [
            "VELOCIDADE DE CONEXÃO",
            "FACILIDADE DE ACESSO À REDE",
        ],
        "SANITÁRIOS": [
            "QUANTIDADE DE BANHEIROS",
            "LIMPEZA DOS BANHEIROS",
            "MANUTENÇÃO GERAL DOS SANITÁRIOS",
        ],
        "PROCESSO DE RESTITUIÇÃO DE BAGAGENS": [
            "FACILIDADE DE IDENTIFICAÇÃO DA ESTEIRA DE RESTITUIÇÃO",
            "TEMPO DE RESTITUIÇÃO",
            "INTEGRIDADE DA BAGAGEM",
        ],
    }

    for parent, children in parent_to_children.items():
        mark_not_applicable(df, children, is_rating_leq_3(df, parent))
        # По методологии SAC субиндикаторы задаются только если
        # макроиндикатор получил оценку 1, 2 или 3. Поэтому при 1-3
        # дочерние вопросы применимы, а при 4-5 или отсутствии оценки
        # помечаются как Not applicable.

    return df


def create_target(df: pd.DataFrame) -> pd.DataFrame:
    # Бинарная целевая переменная: пассажир считается liked, если общая
    # удовлетворенность равна 4 или 5. Оценки 1-3 попадают в класс 0.
    # Строки без числовой оценки удовлетворенности удаляются: иначе NaN
    # через isin([4, 5]) ошибочно превращался бы в класс 0.
    df = df.copy()
    satisfaction = score_to_numeric(df[SATISFACTION_COLUMN])
    # satisfaction содержит числовые оценки 1-5, а отсутствующие/NS становятся NaN.
    df = df.loc[satisfaction.notna()].copy()
    # loc оставляет только строки, где satisfaction не NaN. copy нужен, чтобы
    # pandas не выдавал SettingWithCopyWarning при создании target.
    satisfaction = satisfaction.loc[df.index]
    # После фильтрации df синхронизируем satisfaction по тем же индексам.
    df[TARGET_COLUMN] = satisfaction.isin([4, 5]).astype(int)
    # isin([4, 5]) дает True для довольных; astype(int) переводит True/False в 1/0.
    return df


def normalize_selected_high_cardinality_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Обрабатывает самые проблемные признаки с большим числом уникальных
    # текстовых вариантов до балансировки и импутации.
    df = df.copy()
    if "MOTIVO DA VIAGEM" in df.columns:
        df["MOTIVO DA VIAGEM"] = df["MOTIVO DA VIAGEM"].map(normalize_trip_purpose)
        # map применяет normalize_trip_purpose к каждому ответу о цели поездки.
    if "MEIO DE  TRANSPORTE PARA O AEROPORTO" in df.columns:
        df["MEIO DE  TRANSPORTE PARA O AEROPORTO"] = (
            df["MEIO DE  TRANSPORTE PARA O AEROPORTO"]
            .map(normalize_transport_to_airport)
        )
    for col in ["ANTECEDÊNCIA", "TEMPO DE ESPERA"]:
        if col in df.columns:
            df[col] = df[col].map(normalize_time_range_to_seconds)
            # Обе колонки содержат диапазоны времени; после map становятся
            # числовыми секундами или Not applicable.
    return df


def translate_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Переводит финальные названия колонок на snake_case английские имена.
    # Для созданных индикаторов применимости сохраняет суффикс _is_applicable.
    def translate_column_name(column: str) -> str:
        suffix = " is applicable"
        if column.endswith(suffix):
            # До перевода колонок индикаторы называются "<русская колонка> is applicable".
            # Их надо перевести в "<english_column>_is_applicable".
            base = column[: -len(suffix)]
            # Срез убирает текстовый суффикс " is applicable" с конца строки.
            translated_base = COLUMN_TRANSLATIONS.get(base, base)
            return f"{translated_base}_is_applicable"
        return COLUMN_TRANSLATIONS.get(column, column)

    df = df.copy()
    df.columns = [translate_column_name(col) for col in df.columns]
    # Список новой длины равен числу старых колонок; pandas заменяет все имена.
    return df


def is_numeric_like(series: pd.Series, threshold=0.90) -> bool:
    # Определяет, можно ли колонку считать числовой, игнорируя специальные
    # ответы вроде Not applicable и NS/NR. Это важно для рейтинговых колонок,
    # где большинство значений числовые, но часть строк содержит служебные
    # текстовые ответы.
    non_missing = series.dropna()
    # dropna удаляет обычные NaN, но не удаляет строковые "NS/NR" или Not applicable.
    non_missing = non_missing[
        ~non_missing.astype(str).str.strip().isin(NON_NUMERIC_MISSING_VALUES)
    ]
    # astype(str).str.strip() приводит значения к строкам без пробелов.
    # isin(...) ищет служебные нечисловые значения.
    # ~ инвертирует маску, поэтому остаются только потенциально содержательные числа.
    if non_missing.empty:
        return False
    numeric = pd.to_numeric(non_missing, errors="coerce")
    # После coerce все нечисловые значения становятся NaN.
    # numeric.notna().mean() - доля значений, которые удалось прочитать как числа.
    # threshold=0.90 означает: колонка считается числовой, если >=90% значений числовые.
    return numeric.notna().mean() >= threshold


def impute_remaining_missing(df: pd.DataFrame) -> pd.DataFrame:
    # Финальная обработка пропусков:
    # 1. Делит признаки на числовые и категориальные.
    # 2. Для числовых колонок с Not applicable создает *_is_applicable.
    # 3. Переводит числовые признаки в numeric, оставляя NaN для
    # дальнейшей импутации внутри sklearn Pipeline.
    # Такой порядок сохраняет информацию о структурной неприменимости
    # без расчета медиан/мод на всем датасете до train/test split.
    df = df.copy()

    feature_cols = [col for col in df.columns if col != TARGET_COLUMN]
    # Все колонки, кроме target, являются признаками. Target нельзя имьютировать
    # и нельзя случайно включать в списки признаков.
    numeric_cols = [
        col for col in feature_cols
        if is_numeric_like(df[col])
    ]
    # Числовые колонки определяются не по dtype, а по содержимому: после чтения
    # CSV/XLSX многие числовые рейтинги приходят как object из-за NS/NR.
    categorical_cols = [
        col for col in feature_cols
        if col not in numeric_cols
    ]
    # Все признаки, которые не распознаны как числовые, считаются категориальными.

    numeric_not_applicable_cols = [
        col for col in numeric_cols
        if df[col].eq(NOT_APPLICABLE).any()
    ]
    # eq(NOT_APPLICABLE) делает поэлементное сравнение с "Not applicable".
    # any() проверяет, есть ли хотя бы одно такое значение в колонке.

    applicability_indicators = pd.DataFrame(
        {
            f"{col} is applicable": df[col].ne(NOT_APPLICABLE).astype(int)
            for col in numeric_not_applicable_cols
        },
        index=df.index,
    )
    # Для каждой числовой колонки с Not applicable создается индикатор:
    # 1 = вопрос был применим, 0 = вопрос не относился к этой строке.
    # ne означает "not equal"; astype(int) переводит True/False в 1/0.
    # index=df.index гарантирует, что новые колонки выровнены по тем же строкам.

    if len(applicability_indicators.columns) > 0:
        df = pd.concat([df, applicability_indicators], axis=1)
        # axis=1 означает склеивание по колонкам, а не добавление строк.
        numeric_cols = numeric_cols + applicability_indicators.columns.tolist()
        # Индикаторы 0/1 тоже числовые, поэтому добавляем их в numeric_cols,
        # чтобы дальше они не попали в categorical_cols.

    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        # Все numeric_cols принудительно переводятся в числа. Not applicable,
        # NS/NR и другие нечисловые остатки становятся NaN. Эти пропуски
        # заполняются позже внутри sklearn Pipeline, отдельно для каждого
        # train fold или train split.

    return df, numeric_cols, categorical_cols, numeric_not_applicable_cols


def load_all_resources() -> pd.DataFrame:
    # Загружает все квартальные CSV/XLSX/XLSM из SOURCE_DIR и объединяет
    # их в один DataFrame. Колонка source_file добавляется только для
    # трассировки происхождения строк и затем удаляется перед обучением.
    files = sorted(
        list(SOURCE_DIR.glob("*.csv"))
        + list(SOURCE_DIR.glob("*.xlsx"))
        + list(SOURCE_DIR.glob("*.xlsm"))
    )
    # glob("*.csv") ищет файлы нужного расширения в SOURCE_DIR.
    # list(...) нужен, чтобы сложить результаты разных glob.
    # sorted(...) фиксирует порядок чтения файлов для воспроизводимости.
    if not files:
        raise FileNotFoundError(f"No CSV/XLSX files found in {SOURCE_DIR}")

    frames = []
    for path in files:
        print(f"reading {path.name}")
        df = read_resource(path)
        df = clean_empty_values(df)
        df["source_file"] = path.name
        # source_file временно сохраняет имя исходного файла для отладки.
        # Перед сохранением итогового датасета эта колонка удаляется через DROP_COLUMNS.
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True, sort=False)
    # ignore_index=True создает новый сквозной индекс 0..N-1 вместо сохранения
    # индексов из отдельных файлов.
    # sort=False не сортирует имена колонок, поэтому порядок ближе к исходным файлам.
    combined = combined.rename(columns=COLUMN_RENAMES)
    # rename применяет ручные исправления имен колонок уже после объединения.
    return combined


def balance_classes(df: pd.DataFrame) -> pd.DataFrame:
    # В исходных данных класс liked сильно преобладает. Для честного обучения
    # и оценки модели делаем undersampling большинства до размера меньшинства,
    # затем перемешиваем строки. Это дает баланс 50/50.
    counts = df[TARGET_COLUMN].value_counts()
    # value_counts считает количество строк каждого класса target.
    if len(counts) != 2:
        raise ValueError(f"Expected two target classes, got: {counts.to_dict()}")

    min_count = int(counts.min())
    # min_count - размер меньшего класса. До него уменьшаем больший класс.
    balanced = (
        df
        .groupby(TARGET_COLUMN, group_keys=False)
        # groupby(TARGET_COLUMN) делит данные на группы liked=0 и liked=1.
        # group_keys=False не добавляет значение группы в индекс результата.
        .sample(n=min_count, random_state=RANDOM_STATE)
        # sample(n=min_count) случайно берет одинаковое число строк из каждого класса.
        # random_state делает выбор строк повторяемым.
        .sample(frac=1.0, random_state=RANDOM_STATE)
        # sample(frac=1.0) перемешивает 100% строк без изменения размера датасета.
        .reset_index(drop=True)
        # reset_index(drop=True) создает новый индекс 0..N-1 и удаляет старый.
    )
    return balanced


def finalize_dataset(df: pd.DataFrame):
    # Общая финальная обработка для full и balanced версий датасета.
    # Важно прогонять обе версии через один и тот же набор шагов, чтобы
    # имена колонок, переводы, типы и *_is_applicable создавались одинаково.
    (
        df,
        numeric_cols,
        categorical_cols,
        numeric_not_applicable_cols,
    ) = impute_remaining_missing(df)
    df = translate_output_columns(df)
    df = normalize_final_categorical_values(df)
    df = translate_remaining_text_values(df)

    numeric_cols = [COLUMN_TRANSLATIONS.get(col, col) for col in numeric_cols]
    categorical_cols = [COLUMN_TRANSLATIONS.get(col, col) for col in categorical_cols]
    numeric_not_applicable_cols = [
        COLUMN_TRANSLATIONS.get(col, col)
        for col in numeric_not_applicable_cols
    ]

    return df, numeric_cols, categorical_cols, numeric_not_applicable_cols


def main() -> None:
    # Основной пайплайн сборки датасета. Порядок шагов важен:
    # сначала читаем и нормализуем исходные ответы, затем создаем target,
    # затем восстанавливаем маршрутизацию анкеты, балансируем классы,
    # после этого импутируем пропуски и переводим имена/значения на английский.
    OUTPUT_DIR.mkdir(exist_ok=True)
    # mkdir создает выходную папку. exist_ok=True не падает, если папка уже есть.

    df = load_all_resources()
    # Читает и объединяет все исходные CSV/XLSX/XLSM.
    df = create_target(df)
    # Удаляет строки без числовой SATISFAÇÃO GERAL и создает liked.
    df = apply_questionnaire_routing(df)
    # Проставляет Not applicable там, где вопрос не должен был задаваться.
    df = normalize_selected_high_cardinality_columns(df)
    # Уменьшает кардинальность свободного текста и переводит время в секунды.
    df = df.drop(columns=[col for col in DROP_COLUMNS if col in df.columns])
    # drop(columns=...) удаляет только реально существующие колонки.
    # List comprehension защищает от ошибки, если в каком-то наборе данных
    # часть колонок отсутствует.

    full_counts = df[TARGET_COLUMN].value_counts().sort_index()
    # Это распределение классов в полной, несбалансированной версии датасета.
    full_df = df.copy()
    # copy сохраняет full-версию до undersampling, чтобы потом записать
    # passenger_survey_full.csv.

    before_balance_counts = full_counts
    # sort_index сортирует классы как 0, 1, чтобы отчет был стабильным.
    balanced_df = balance_classes(df)
    after_balance_counts = balanced_df[TARGET_COLUMN].value_counts().sort_index()

    full_df, full_numeric_cols, full_categorical_cols, full_numeric_na_cols = (
        finalize_dataset(full_df)
    )

    (
        balanced_df,
        numeric_cols,
        categorical_cols,
        numeric_not_applicable_cols,
    ) = finalize_dataset(balanced_df)

    full_output_csv = OUTPUT_DIR / "passenger_survey_full.csv"
    output_csv = OUTPUT_DIR / "passenger_survey_balanced.csv"
    report_json = OUTPUT_DIR / "passenger_survey_balanced_report.json"

    full_df.to_csv(full_output_csv, index=False, encoding="utf-8-sig")
    balanced_df.to_csv(output_csv, index=False, encoding="utf-8-sig")
    # index=False не сохраняет pandas index как отдельную колонку.
    # utf-8-sig добавляет BOM, чтобы Excel корректно открывал UTF-8 CSV.

    report = {
        "source_dir": str(SOURCE_DIR),
        "output_csv": str(output_csv),
        "full_output_csv": str(full_output_csv),
        "n_rows": int(len(balanced_df)),
        "n_columns": int(len(balanced_df.columns)),
        "full_n_rows": int(len(full_df)),
        "full_n_columns": int(len(full_df.columns)),
        "class_counts_before_balance": {
            str(k): int(v) for k, v in before_balance_counts.to_dict().items()
        },
        "class_counts_full": {
            str(k): int(v) for k, v in full_counts.to_dict().items()
        },
        # k переводится в str, потому что JSON-ключи должны быть строками.
        # v переводится в int, потому что pandas/numpy int64 не всегда сериализуется JSON.
        "class_counts_after_balance": {
            str(k): int(v) for k, v in after_balance_counts.to_dict().items()
        },
        "dropped_columns": [col for col in DROP_COLUMNS if col != TARGET_COLUMN],
        "numeric_imputed_columns": numeric_cols,
        "categorical_imputed_columns": categorical_cols,
        "numeric_columns_with_not_applicable_indicator": numeric_not_applicable_cols,
        "full_numeric_imputed_columns": full_numeric_cols,
        "full_categorical_imputed_columns": full_categorical_cols,
        "full_numeric_columns_with_not_applicable_indicator": full_numeric_na_cols,
        "created_applicability_indicator_columns": [
            f"{col}_is_applicable"
            for col in numeric_not_applicable_cols
        ],
        # Это финальные имена созданных индикаторов применимости после перевода колонок.
        "not_applicable_value": NOT_APPLICABLE,
        "random_state": RANDOM_STATE,
        "balancing": "undersample majority class to minority class size",
    }
    report_json.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    # ensure_ascii=False сохраняет русские/португальские символы читаемыми.
    # indent=2 делает JSON удобным для просмотра человеком.

    print(f"saved full: {full_output_csv}")
    print(f"saved: {output_csv}")
    print(f"report: {report_json}")
    print(f"full rows: {len(full_df)}")
    print(f"balanced rows: {len(balanced_df)}")
    print(f"class counts before balance: {before_balance_counts.to_dict()}")
    print(f"class counts after balance: {after_balance_counts.to_dict()}")


if __name__ == "__main__":
    main()
