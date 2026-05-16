import csv
import hashlib
import json
import re
import shutil
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


PACKAGE_API_URL = (
    "https://dados.transportes.gov.br/api/3/action/package_show"
    "?id=pesquisa-de-satisfacao-do-passageiro-em-aeroportos"
)

ROOT = Path(__file__).resolve().parent
DOWNLOAD_DIR = ROOT / "passenger_survey_downloads"
GROUPS_DIR = ROOT / "passenger_survey_column_groups"
REPORT_PATH = GROUPS_DIR / "column_group_report.json"


def safe_filename(name: str, url: str) -> str:
    url_name = url.rstrip("/").split("/")[-1]
    suffix = Path(url_name).suffix.lower()
    base = Path(url_name).stem or name
    base = re.sub(r"[^A-Za-z0-9._-]+", "-", base).strip("-")
    if not base:
        base = "resource"
    return f"{base}{suffix}"


def fetch_json(url: str) -> dict:
    with urllib.request.urlopen(url, timeout=60) as response:
        return json.load(response)


def download(url: str, path: Path) -> None:
    if path.exists() and path.stat().st_size > 0:
        return
    local_copy = ROOT / path.name
    if local_copy.exists() and local_copy.stat().st_size > 0:
        shutil.copy2(local_copy, path)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    last_error = None
    for attempt in range(1, 5):
        try:
            with urllib.request.urlopen(request, timeout=180) as response:
                path.write_bytes(response.read())
            return
        except Exception as error:
            last_error = error
            if attempt == 4:
                break
            time.sleep(attempt * 2)
    raise last_error


def normalize_header_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_xlsx_columns(path: Path) -> tuple[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [name for name in wb.sheetnames if name.upper() != "LEGENDA"]
    ws = wb[sheet_names[0]]

    best_row = None
    best_score = -1
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [normalize_header_cell(value) for value in row]
        score = sum(bool(value) for value in values)
        if "CHAVE" in values:
            best_row = values
            break
        if score > best_score:
            best_row = values
            best_score = score
        if row_index >= 10:
            break

    columns = [value for value in best_row if value]
    return ws.title, columns


def decode_csv_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin1", errors="replace")


def extract_csv_columns(path: Path) -> tuple[str, list[str]]:
    text = decode_csv_bytes(path.read_bytes())
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(text.splitlines(), dialect)
    best_row = None
    best_score = -1
    for row_index, row in enumerate(reader, start=1):
        values = [normalize_header_cell(value) for value in row]
        score = sum(bool(value) for value in values)
        if "CHAVE" in values:
            best_row = values
            break
        if score > best_score:
            best_row = values
            best_score = score
        if row_index >= 10:
            break

    columns = [value for value in best_row if value]
    return "csv", columns


def extract_columns(path: Path) -> tuple[str, list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return extract_xlsx_columns(path)
    if suffix == ".csv":
        return extract_csv_columns(path)
    raise ValueError(f"Unsupported file type: {path}")


def schema_key(columns: list[str]) -> str:
    payload = json.dumps(columns, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:12]


def main() -> None:
    package = fetch_json(PACKAGE_API_URL)
    resources = []
    for resource in package["result"]["resources"]:
        url = resource.get("url") or ""
        fmt = (resource.get("format") or "").lower().strip(".")
        suffix = Path(url.split("?", 1)[0]).suffix.lower().strip(".")
        if fmt in {"csv", "xlsx", "xls"} or suffix in {"csv", "xlsx", "xls"}:
            resources.append(resource)

    DOWNLOAD_DIR.mkdir(exist_ok=True)
    GROUPS_DIR.mkdir(exist_ok=True)

    groups = {}
    resource_rows = []

    for index, resource in enumerate(resources, start=1):
        name = resource.get("name") or f"resource-{index}"
        url = resource["url"]
        filename = safe_filename(name, url)
        path = DOWNLOAD_DIR / filename

        print(f"[{index}/{len(resources)}] downloading/checking {filename}")
        download(url, path)

        sheet, columns = extract_columns(path)
        key = schema_key(columns)
        groups.setdefault(key, {"columns": columns, "files": []})
        groups[key]["files"].append(filename)

        resource_rows.append(
            {
                "file": filename,
                "name": name,
                "format": resource.get("format"),
                "url": url,
                "sheet": sheet,
                "n_columns": len(columns),
                "schema_key": key,
                "columns": columns,
            }
        )

    for old_group in GROUPS_DIR.glob("group_*"):
        if old_group.is_dir():
            shutil.rmtree(old_group)

    sorted_groups = sorted(
        groups.items(),
        key=lambda item: (-len(item[1]["files"]), item[0]),
    )

    report_groups = []
    for group_index, (key, group) in enumerate(sorted_groups, start=1):
        group_dir = GROUPS_DIR / f"group_{group_index:02d}_{key}"
        group_dir.mkdir(parents=True, exist_ok=True)
        for filename in group["files"]:
            shutil.copy2(DOWNLOAD_DIR / filename, group_dir / filename)
        (group_dir / "columns.txt").write_text(
            "\n".join(group["columns"]) + "\n",
            encoding="utf-8",
        )
        report_groups.append(
            {
                "group": group_dir.name,
                "schema_key": key,
                "n_files": len(group["files"]),
                "n_columns": len(group["columns"]),
                "files": sorted(group["files"]),
                "columns": group["columns"],
            }
        )

    report = {
        "source": PACKAGE_API_URL,
        "n_resources": len(resources),
        "n_column_groups": len(report_groups),
        "groups": report_groups,
        "resources": resource_rows,
    }
    REPORT_PATH.write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"resources: {len(resources)}")
    print(f"column groups: {len(report_groups)}")
    for group in report_groups:
        print(
            f"{group['group']}: {group['n_files']} files, "
            f"{group['n_columns']} columns"
        )
        for filename in group["files"]:
            print(f"  - {filename}")
    print(f"report: {REPORT_PATH}")


if __name__ == "__main__":
    main()
